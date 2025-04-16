import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Load predictions and labels
pred_df = pd.read_csv('./human_identification/predictions.csv')
label_df = pd.read_csv('./human_identification/test/labels.csv')

# Fix: extract index from first (unnamed) column
pred_df['index'] = pred_df.iloc[:, 0].str.extract(r'(\d+)').astype(int)

# Merge predictions with true labels
merged_df = pd.merge(pred_df, label_df, on='index', how='inner')
merged_df['correct'] = merged_df['prediction'] == merged_df['class']

# Save with formatting
output_excel_path = './human_identification/comparison.xlsx'
wb = Workbook()
ws = wb.active
ws.title = "Results"

# Write headers and rows
for r in dataframe_to_rows(merged_df, index=False, header=True):
    ws.append(r)

# Color rows
green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    correct_cell = row[merged_df.columns.get_loc('correct')]
    color = green if correct_cell.value else red
    for cell in row:
        cell.fill = color

wb.save(output_excel_path)
print(f"Comparison saved as Excel to {output_excel_path}")
